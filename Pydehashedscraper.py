#!/usr/bin/env python
import requests
import openpyxl
import re
from openpyxl import Workbook
from getpass import getpass

def dehashed_search(api_key, email, domain):
    query = f"domain:{domain}"
    url = f"https://api.dehashed.com/search?query={query}"
    headers = {'Accept': 'application/json'}

    response = requests.get(url, auth=(email, api_key), headers=headers)
    response.raise_for_status()

    return response.json()

def save_to_excel(data, file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dehashed Data"
    
    headers = ["Email", "Password", "Hashed Password", "Database Name"]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header
    
    for row, entry in enumerate(data['entries'], start=2):
        ws.cell(row=row, column=1).value = entry.get('email', '')
        ws.cell(row=row, column=2).value = entry.get('password', '')
        ws.cell(row=row, column=3).value = entry.get('hashed_password', '')
        ws.cell(row=row, column=4).value = entry.get('database_name', '')
    
    wb.save(file_name)

def is_valid_email(email):
    regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return re.match(regex, email)

if __name__ == "__main__":
    email = input("Enter your DeHashed email address: ")
    while not is_valid_email(email):
        print("Invalid email address. Please try again.")
        email = input("Enter your DeHashed email address: ")

    api_key = getpass("Enter your DeHashed API key: ")
    while len(api_key) == 0:
        print("API key cannot be empty. Please try again.")
        api_key = getpass("Enter your DeHashed API key: ")

    domain = input("Enter the domain to search for: ")
    while len(domain) == 0:
        print("Domain cannot be empty. Please try again.")
        domain = input("Enter the domain to search for: ")

    file_name = input("Enter the output Excel file name (e.g. output.xlsx): ")
    while not file_name.endswith('.xlsx'):
        print("Invalid file name. Please provide a file name with .xlsx extension.")
        file_name = input("Enter the output Excel file name (e.g. output.xlsx): ")

    data = dehashed_search(api_key, email, domain)
    save_to_excel(data, file_name)

    print(f"Data saved to {file_name}")
